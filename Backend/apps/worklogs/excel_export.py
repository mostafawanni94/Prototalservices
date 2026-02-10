"""
Excel Export for Customer - Template-Based Export using Master.xlsx

This module implements the high-fidelity Excel export following the
Template-Based Restoration Pattern from the Audit Reporting Standards.

The Master.xlsx template structure:
- Rows 1-10: Header section with metadata, logos, and column labels
- Row 11+: Data grid for worklog entries
- Sheet: "Overzicht Storm"
"""

import os
from datetime import datetime, time, timedelta
from io import BytesIO
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter



# Path to the Master.xlsx template
TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))),
    'Frontend', 'Master.xlsx'
)


# Dutch day abbreviation mapping
DAY_MAPPING = {
    0: 'ma',  # Monday
    1: 'di',  # Tuesday
    2: 'wo',  # Wednesday
    3: 'do',  # Thursday
    4: 'vr',  # Friday
    5: 'za',  # Saturday
    6: 'zo',  # Sunday
}


def format_time(time_str):
    """Convert time string (HH:MM:SS or HH:MM) to time object."""
    if not time_str:
        return None
    if isinstance(time_str, time):
        return time_str
    try:
        parts = str(time_str).split(':')
        hour = int(parts[0])
        minute = int(parts[1]) if len(parts) > 1 else 0
        return time(hour, minute)
    except (ValueError, IndexError):
        return None


def parse_hours(value):
    """Parse hours value to float."""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(',', '.'))
    except (ValueError, TypeError):
        return 0


def generate_customer_excel_export(
    worklogs,
    customer_name,
    customer_email=None,
    customer_address=None,
    week_from=None,
    week_to=None,
    rayon_name=None,
    medewerker_name=None,
    customer_logo_path=None,
    customer_surcharges=None,
    export_type='hr',  # 'hr' or 'finance'
    hr_email=None
):

    """
    Generate a fresh Excel file with worklog data.
    
    Creates a new Excel file from scratch (single sheet) based on user's
    step-by-step cell specifications, with proper styling.
    
    Args:
        worklogs: List of worklog dictionaries
        customer_name: Name of the customer
        customer_email: Customer email address
        customer_address: Customer address
        week_from: Starting week number
        week_to: Ending week number
        rayon_name: Supervisor/Rayon name
        medewerker_name: Employee name (if filtering by single employee)
        customer_logo_path: Path to customer logo image
        customer_surcharges: List of customer surcharge dicts with name and percentage
    
    Returns:
        BytesIO buffer containing the Excel workbook
    """

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    import os
    
    # Create a new workbook with a single sheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Overzicht'
    
    # === STYLING SETUP ===
    # Fonts - matching Master.xlsx professional look
    title_font = Font(name='Calibri', size=16, bold=True)
    week_number_font = Font(name='Calibri', size=24, bold=True)  # Large week numbers
    header_font = Font(name='Calibri', size=14, bold=True)
    normal_font = Font(name='Calibri', size=11)
    label_font = Font(name='Calibri', size=12, bold=True)
    company_font = Font(name='Calibri', size=14, bold=True)
    rayon_value_font = Font(name='Calibri', size=14, bold=True)
    
    # Fills - background colors
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')  # Light green
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Hide gridlines for cleaner look
    ws.sheet_view.showGridLines = False
    
    # Set column widths (matching Master.xlsx proportions)
    column_widths = {
        'A': 22,   # Labels column (Email:, Adres, Week, Rayon)
        'B': 5,
        'C': 5,
        'D': 40,   # Values column - wider
        'E': 12,   # Week to
        'F': 12,   
        'G': 12,
        'H': 12,
        'I': 28,   # Customer name column
        'J': 12,
        'K': 12,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Set row heights (taller for professional look)
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 35  # Week row - taller for large numbers
    ws.row_dimensions[6].height = 8   # Empty row - very small
    ws.row_dimensions[7].height = 30  # Rayon row - taller
    
    # === HEADER SECTION (Rows 1-7) ===
    
    # Row 1: D = Title with date
    today = datetime.now().strftime('%d-%m-%Y')
    ws['D1'] = f'Inleenovereenkomst per {today}'
    ws['D1'].font = title_font
    
    # Row 2: A = Customer name with "B.V." suffix (logo next to it)
    ws['A2'] = f'{customer_name} B.V.' if customer_name else ''
    ws['A2'].font = company_font
    
    # Row 3: A = "Email:", D = Email value (use HR email if available)
    ws['A3'] = 'Email:'
    ws['A3'].font = label_font
    display_email = hr_email if hr_email else customer_email
    if display_email:
        ws['D3'] = display_email
        ws['D3'].font = normal_font

    # I3 is now empty (deleted)

    
    # Row 4: A = "Adres", D = Address value
    ws['A4'] = 'Adres'
    ws['A4'].font = label_font
    if customer_address:
        ws['D4'] = customer_address
        ws['D4'].font = normal_font
    
    # Row 5: A = "Week", D = week from, E = week to (LARGE numbers)
    ws['A5'] = 'Week'
    ws['A5'].font = label_font
    if week_from is not None:
        ws['D5'] = week_from
        ws['D5'].font = week_number_font
        ws['D5'].alignment = Alignment(horizontal='center', vertical='center')
    if week_to is not None:
        ws['E5'] = week_to
        ws['E5'].font = week_number_font
        ws['E5'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Row 6: Empty
    
    # Row 7: A = "Rayon", D = Rayon value with GREEN background
    ws['A7'] = 'Rayon'
    ws['A7'].font = label_font
    if rayon_name:
        ws['D7'] = f'Medewerker {rayon_name}'
        ws['D7'].font = rayon_value_font
        ws['D7'].fill = green_fill
        ws['D7'].border = thin_border
        ws['D7'].alignment = Alignment(horizontal='center', vertical='center')
        # Extend green fill to E7 as well (merged look)
        ws['E7'].fill = green_fill
        ws['E7'].border = thin_border
    
    # === INSERT CUSTOMER LOGO (right next to company name in A2) ===
    if customer_logo_path and os.path.exists(customer_logo_path):
        try:
            img = XLImage(customer_logo_path)
            # Scale image to fit nicely next to company name
            img.width = 80
            img.height = 40
            ws.add_image(img, 'C2')  # Position right next to company name
        except Exception as e:
            # If logo fails to load, continue without it
            pass


    
    # === TABLE SECTION (One continuous table from A-X) ===
    
    # Row 8: Customer name header above hour columns + Storm logo position
    ws['L8'] = customer_name or 'Customer'
    ws['L8'].font = header_font
    ws['L8'].alignment = Alignment(horizontal='center')
    
    # Column layout differs between HR and Finance exports
    # HR: J-Q for hour breakdown, R for TOTAAL UREN
    # Finance: J for Totaal Uren, K-R for breakdown, S for TOTAAL BEDRAG
    
    if export_type == 'hr':
        row9_headers = {
            'A': 'Medewerker naam',
            'B': 'FUNCTEI',
            'C': 'DAG',
            'D': 'DATUM',
            'E': 'BEGIN PROJECT',
            'F': 'EINDE PROJECT',
            'G': 'pauze',
            'H': 'Pauza Begin',
            'I': 'Pauza Einde',
            'J': 'Regulieredienst',           # Total hours if ONLY normal hours worked, else 0
            'K': 'Ploegendiensturen',         # Total hours if ANY surcharge hours worked, else 0
            'L': 'Normaal Uren',              # Normal weekday hours (breakdown)
            'M': 'Night Shift',
            'N': 'Zaterdag en zondagen',
            'O': 'Feestdagen',
            'P': 'overuur',
            'Q': 'Slaap uren',
            'R': 'Toeslagen',
            'S': 'TOTAAL UREN',
        }
        hour_columns = ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']
        last_col = 'S'
    else:  # finance
        row9_headers = {
            'A': 'Medewerker naam',
            'B': 'FUNCTEI',
            'C': 'DAG',
            'D': 'DATUM',
            'E': 'BEGIN PROJECT',
            'F': 'EINDE PROJECT',
            'G': 'pauze',
            'H': 'Pauza Begin',
            'I': 'Pauza Einde',
            'J': 'Totaal Uren',
            'K': 'Regulieredienst',           # Total hours if ONLY normal hours worked, else 0
            'L': 'Ploegendiensturen',         # Total hours if ANY surcharge hours worked, else 0
            'M': 'Normaal Uren',              # Normal weekday hours (breakdown)
            'N': 'Night Shift',
            'O': 'Zaterdag en zondagen',
            'P': 'Feestdagen',
            'Q': 'overuur',
            'R': 'Slaap uren',
            'S': 'Toeslagen',
            'T': 'TOTAAL BEDRAG',
        }
        hour_columns = ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']
        last_col = 'T'

    for col, header in row9_headers.items():
        ws[f'{col}9'] = header
        ws[f'{col}9'].font = header_font
        ws[f'{col}9'].border = thin_border
        ws[f'{col}9'].alignment = Alignment(horizontal='center', wrap_text=True)
        # Green fill for hour type columns
        if col in hour_columns:
            ws[f'{col}9'].fill = green_fill

    
    # Row 10: Percentages under hour type columns
    if export_type == 'hr':
        default_hour_percentages = {
            'J': '',       # Regulieredienst - no percentage
            'K': '',       # Ploegendiensturen - no percentage
            'L': '100%',   # Normaal Uren
            'M': '120%',   # Night shift
            'N': '130%',   # Weekend
            'O': '200%',   # Holiday
            'P': '110%',   # Overtime
            'Q': '90%',    # Sleep hours
            'R': 'EPZ',    # Allowance placeholder
        }
    else:  # finance
        default_hour_percentages = {
            'J': '',       # Totaal Uren - no percentage
            'K': '',       # Regulieredienst - no percentage
            'L': '',       # Ploegendiensturen - no percentage
            'M': '100%',   # Normaal Uren
            'N': '120%',   # Night shift
            'O': '130%',   # Weekend
            'P': '200%',   # Holiday
            'Q': '110%',   # Overtime
            'R': '90%',    # Sleep hours
            'S': 'EPZ',    # Allowance placeholder
        }
    
    # Override with customer surcharges if provided
    if customer_surcharges:
        # Column mapping differs based on export type
        night_col = 'M' if export_type == 'hr' else 'N'
        weekend_col = 'N' if export_type == 'hr' else 'O'
        holiday_col = 'O' if export_type == 'hr' else 'P'
        overtime_col = 'P' if export_type == 'hr' else 'Q'
        sleep_col = 'Q' if export_type == 'hr' else 'R'
        
        for sc in customer_surcharges:
            name = sc.get('name', '').lower()
            pct = sc.get('percentage', 0)
            if 'night' in name or 'nacht' in name:
                default_hour_percentages[night_col] = f"{int(pct) + 100}%"
            elif 'zat' in name or 'zon' in name or 'weekend' in name:
                default_hour_percentages[weekend_col] = f"{int(pct) + 100}%"
            elif 'feest' in name or 'holiday' in name:
                default_hour_percentages[holiday_col] = f"{int(pct) + 100}%"
            elif 'over' in name:
                default_hour_percentages[overtime_col] = f"{int(pct) + 100}%"
            elif 'slaap' in name or 'sleep' in name:
                default_hour_percentages[sleep_col] = f"{int(pct)}%"
    
    for col, pct in default_hour_percentages.items():
        ws[f'{col}10'] = pct
        ws[f'{col}10'].font = normal_font
        ws[f'{col}10'].fill = green_fill
        ws[f'{col}10'].border = thin_border
        ws[f'{col}10'].alignment = Alignment(horizontal='center')



    
    # Set column widths for continuous table (wider to fit all text)
    if export_type == 'hr':
        table_column_widths = {
            'A': 30, 'B': 28, 'C': 6, 'D': 12, 'E': 14, 'F': 14, 'G': 10,
            'H': 12, 'I': 12,
            'J': 14, 'K': 16, 'L': 16, 'M': 12, 'N': 16, 'O': 12, 'P': 10, 'Q': 12, 'R': 12, 'S': 14,
        }
    else:  # finance
        table_column_widths = {
            'A': 30, 'B': 28, 'C': 6, 'D': 12, 'E': 14, 'F': 14, 'G': 10,
            'H': 12, 'I': 12,
            'J': 12, 'K': 14, 'L': 16, 'M': 16, 'N': 12, 'O': 16, 'P': 12, 'Q': 10, 'R': 12, 'S': 12, 'T': 14,
        }


    
    for col, width in table_column_widths.items():
        ws.column_dimensions[col].width = width
    
    # === DATA ROWS (Starting from Row 11) ===
    data_row = 11
    
    # Dutch day abbreviations
    day_abbrevs = {
        0: 'ma', 1: 'di', 2: 'wo', 3: 'do', 4: 'vr', 5: 'za', 6: 'zo'
    }
    
    for worklog in worklogs:
        # A: Employee name
        ws[f'A{data_row}'] = worklog.get('employee_name', '')
        ws[f'A{data_row}'].border = thin_border
        
        # B: FUNCTEI (Service name)
        ws[f'B{data_row}'] = worklog.get('service_name', '')
        ws[f'B{data_row}'].border = thin_border
        
        # C: DAG (Day abbreviation)
        work_date = worklog.get('work_date')
        if work_date:
            day_abbrev = day_abbrevs.get(work_date.weekday(), '')
            ws[f'C{data_row}'] = day_abbrev
        ws[f'C{data_row}'].border = thin_border
        ws[f'C{data_row}'].alignment = Alignment(horizontal='center')
        
        # D: DATUM (Date)
        if work_date:
            ws[f'D{data_row}'] = work_date.strftime('%d-%b')
        ws[f'D{data_row}'].border = thin_border
        ws[f'D{data_row}'].alignment = Alignment(horizontal='center')
        
        # E: BEGIN PROJECT (Start time)
        ws[f'E{data_row}'] = worklog.get('start_time', '')
        ws[f'E{data_row}'].border = thin_border
        ws[f'E{data_row}'].alignment = Alignment(horizontal='center')
        
        # F: EINDE PROJECT (End time)
        ws[f'F{data_row}'] = worklog.get('end_time', '')
        ws[f'F{data_row}'].border = thin_border
        ws[f'F{data_row}'].alignment = Alignment(horizontal='center')
        
        # G: pauze (Break)
        ws[f'G{data_row}'] = worklog.get('break_duration', '0:00')
        ws[f'G{data_row}'].border = thin_border
        ws[f'G{data_row}'].alignment = Alignment(horizontal='center')
        
        # H: Pauza Begin (Break start time) - NEW
        ws[f'H{data_row}'] = worklog.get('break_start', '')
        ws[f'H{data_row}'].border = thin_border
        ws[f'H{data_row}'].alignment = Alignment(horizontal='center')
        
        # I: Pauza Einde (Break end time) - NEW
        ws[f'I{data_row}'] = worklog.get('break_end', '')
        ws[f'I{data_row}'].border = thin_border
        ws[f'I{data_row}'].alignment = Alignment(horizontal='center')
        
        # Hour breakdown columns - layout differs between HR and Finance
        hours_breakdown = worklog.get('hours_breakdown', {})
        calculated_hours = worklog.get('calculated_hours', 0)
        total_hours = float(calculated_hours or 8)  # Default 8 if not calculated
        
        # Extract hour values
        normal_hours = float(hours_breakdown.get('normal_hours', 0) or 0)
        night_hours = float(hours_breakdown.get('night_hours', 0) or 0)
        weekend_hours = float(hours_breakdown.get('weekend_hours', 0) or 0)
        holiday_hours = float(hours_breakdown.get('holiday_hours', 0) or 0)
        overtime_hours = float(hours_breakdown.get('overtime_hours', 0) or 0)
        sleep_hours = float(hours_breakdown.get('sleep_hours', 0) or 0)
        
        # Calculate allowance hours
        allowances_list = worklog.get('allowances', [])
        total_allowance_hours = 0
        if allowances_list and isinstance(allowances_list, list):
            for allowance in allowances_list:
                try:
                    total_allowance_hours += float(allowance.get('hours', 0) or 0)
                except (ValueError, TypeError):
                    pass
        
        # Calculate Regulieredienst and Ploegendiensturen using either/or logic
        # Check if ANY surcharge hours exist (night, weekend, holiday, overtime)
        has_surcharge_hours = (night_hours + weekend_hours + holiday_hours + overtime_hours) > 0
        
        # Either/or logic:
        # - Regulieredienst = total_hours if ONLY normal hours worked (no surcharges), else 0
        # - Ploegendiensturen = total_hours if ANY surcharge hours worked, else 0
        regulieredienst_hours = total_hours if not has_surcharge_hours else 0
        ploegendienst_hours = total_hours if has_surcharge_hours else 0
        
        if export_type == 'hr':
            # HR layout: J=Reguliere, K=Ploegendienst, L=Normaal, M=Night, N=Weekend, O=Holiday, P=Overtime, Q=Sleep, R=Toeslagen, S=TOTAAL
            ws[f'J{data_row}'] = regulieredienst_hours
            ws[f'J{data_row}'].border = thin_border
            ws[f'J{data_row}'].alignment = Alignment(horizontal='center')
            ws[f'J{data_row}'].number_format = '0.00'
            
            ws[f'K{data_row}'] = ploegendienst_hours
            ws[f'K{data_row}'].border = thin_border
            ws[f'K{data_row}'].alignment = Alignment(horizontal='center')
            ws[f'K{data_row}'].number_format = '0.00'
            
            ws[f'L{data_row}'] = normal_hours
            ws[f'L{data_row}'].border = thin_border
            ws[f'L{data_row}'].alignment = Alignment(horizontal='center')
            ws[f'L{data_row}'].number_format = '0.00'
            
            ws[f'M{data_row}'] = night_hours if night_hours else 0
            ws[f'M{data_row}'].border = thin_border
            ws[f'M{data_row}'].number_format = '0.00'
            
            ws[f'N{data_row}'] = weekend_hours if weekend_hours else 0
            ws[f'N{data_row}'].border = thin_border
            ws[f'N{data_row}'].number_format = '0.00'
            
            ws[f'O{data_row}'] = holiday_hours if holiday_hours else 0
            ws[f'O{data_row}'].border = thin_border
            ws[f'O{data_row}'].number_format = '0.00'
            
            ws[f'P{data_row}'] = overtime_hours if overtime_hours else 0
            ws[f'P{data_row}'].border = thin_border
            ws[f'P{data_row}'].number_format = '0.00'
            
            ws[f'Q{data_row}'] = sleep_hours if sleep_hours else 0
            ws[f'Q{data_row}'].border = thin_border
            ws[f'Q{data_row}'].number_format = '0.00'
            
            ws[f'R{data_row}'] = total_allowance_hours
            ws[f'R{data_row}'].border = thin_border
            ws[f'R{data_row}'].number_format = '0.00'
            
            # S: TOTAAL UREN
            ws[f'S{data_row}'] = total_hours
            ws[f'S{data_row}'].border = thin_border
            ws[f'S{data_row}'].alignment = Alignment(horizontal='center')
            ws[f'S{data_row}'].number_format = '0.00'
        
        else:  # finance
            # Finance layout: J=Totaal, K=Reguliere, L=Ploegendienst, M=Normaal, N=Night, O=Weekend, P=Holiday, Q=Overtime, R=Sleep, S=Toeslagen, T=BEDRAG
            ws[f'J{data_row}'] = total_hours
            ws[f'J{data_row}'].border = thin_border
            ws[f'J{data_row}'].alignment = Alignment(horizontal='center')
            ws[f'J{data_row}'].number_format = '0.00'
            
            ws[f'K{data_row}'] = regulieredienst_hours
            ws[f'K{data_row}'].border = thin_border
            ws[f'K{data_row}'].alignment = Alignment(horizontal='center')
            ws[f'K{data_row}'].number_format = '0.00'
            
            ws[f'L{data_row}'] = ploegendienst_hours
            ws[f'L{data_row}'].border = thin_border
            ws[f'L{data_row}'].alignment = Alignment(horizontal='center')
            ws[f'L{data_row}'].number_format = '0.00'
            
            ws[f'M{data_row}'] = normal_hours
            ws[f'M{data_row}'].border = thin_border
            ws[f'M{data_row}'].alignment = Alignment(horizontal='center')
            ws[f'M{data_row}'].number_format = '0.00'
            
            ws[f'N{data_row}'] = night_hours if night_hours else 0
            ws[f'N{data_row}'].border = thin_border
            ws[f'N{data_row}'].number_format = '0.00'
            
            ws[f'O{data_row}'] = weekend_hours if weekend_hours else 0
            ws[f'O{data_row}'].border = thin_border
            ws[f'O{data_row}'].number_format = '0.00'
            
            ws[f'P{data_row}'] = holiday_hours if holiday_hours else 0
            ws[f'P{data_row}'].border = thin_border
            ws[f'P{data_row}'].number_format = '0.00'
            
            ws[f'Q{data_row}'] = overtime_hours if overtime_hours else 0
            ws[f'Q{data_row}'].border = thin_border
            ws[f'Q{data_row}'].number_format = '0.00'
            
            ws[f'R{data_row}'] = sleep_hours if sleep_hours else 0
            ws[f'R{data_row}'].border = thin_border
            ws[f'R{data_row}'].number_format = '0.00'
            
            ws[f'S{data_row}'] = total_allowance_hours
            ws[f'S{data_row}'].border = thin_border
            ws[f'S{data_row}'].number_format = '0.00'
            
            # T: TOTAAL BEDRAG (currency)
            total_amount = float(worklog.get('calculated_price', 0))
            ws[f'T{data_row}'] = total_amount
            ws[f'T{data_row}'].border = thin_border
            ws[f'T{data_row}'].alignment = Alignment(horizontal='right')
            ws[f'T{data_row}'].number_format = '€#,##0.00'




        
        data_row += 1

    
    # === Add TOTAL row at the end ===
    total_row = data_row
    
    # "Total" label in column A
    ws[f'A{total_row}'] = 'Total'
    ws[f'A{total_row}'].font = Font(bold=True, size=11)
    ws[f'A{total_row}'].border = thin_border
    
    # Empty cells for columns B-I (including break time columns H, I which don't need totals)
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws[f'{col}{total_row}'] = ''
        ws[f'{col}{total_row}'].border = thin_border
    
    # Start row for data (row 11 is first data row)
    first_data_row = 11
    last_data_row = total_row - 1
    
    if export_type == 'hr':
        # HR Export: Sum hour breakdown columns (J-R) and TOTAAL UREN (S)
        for col in ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']:
            ws[f'{col}{total_row}'] = f'=SUM({col}{first_data_row}:{col}{last_data_row})'
            ws[f'{col}{total_row}'].font = Font(bold=True, size=11)
            ws[f'{col}{total_row}'].border = thin_border
            ws[f'{col}{total_row}'].alignment = Alignment(horizontal='center')
            ws[f'{col}{total_row}'].fill = green_fill
            ws[f'{col}{total_row}'].number_format = '#,##0.00'
    
    elif export_type == 'finance':
        # Finance Export: Sum hour columns (J-S) and TOTAAL BEDRAG (T)
        for col in ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']:
            ws[f'{col}{total_row}'] = f'=SUM({col}{first_data_row}:{col}{last_data_row})'
            ws[f'{col}{total_row}'].font = Font(bold=True, size=11)
            ws[f'{col}{total_row}'].border = thin_border
            ws[f'{col}{total_row}'].alignment = Alignment(horizontal='center')
            ws[f'{col}{total_row}'].fill = green_fill
            ws[f'{col}{total_row}'].number_format = '#,##0.00'
        
        # TOTAAL BEDRAG sum (T) - currency format
        ws[f'T{total_row}'] = f'=SUM(T{first_data_row}:T{last_data_row})'
        ws[f'T{total_row}'].font = Font(bold=True, size=11)
        ws[f'T{total_row}'].border = thin_border
        ws[f'T{total_row}'].alignment = Alignment(horizontal='right')
        ws[f'T{total_row}'].fill = green_fill
        ws[f'T{total_row}'].number_format = '€#,##0.00'



    
    # === Save to buffer and return ===
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    
    return buffer







@api_view(['GET'])
@permission_classes([IsAuthenticated])
def excel_export_view(request):
    """
    Django view to handle Excel export requests.
    
    Expected query parameters:
        - customer_id: UUID of the customer
        - week_start: Start week in format YYYY-Www (e.g., 2025-W49)
        - week_end: End week in format YYYY-Www
        - employee_ids: Comma-separated list of employee UUIDs (optional)
        - supervisor_id: UUID of supervisor/rayon (optional)
    """

    from apps.worklogs.models import WorkEntry
    from apps.customers.models import Customer
    
    # Parse request parameters
    customer_id = request.GET.get('customer_id')
    week_start = request.GET.get('week_start')
    week_end = request.GET.get('week_end')
    employee_ids = request.GET.get('employee_ids', '').split(',') if request.GET.get('employee_ids') else []
    supervisor_id = request.GET.get('supervisor_id')
    export_type = request.GET.get('export_type', 'hr')  # 'hr' or 'finance'
    
    # Validate customer
    try:
        customer = Customer.objects.prefetch_related('contacts').get(id=customer_id)
        customer_name = customer.company_name
        # Get primary email from contacts (General Manager)
        email_contact = customer.contacts.filter(contact_type='email', is_primary=True).first()
        if not email_contact:
            email_contact = customer.contacts.filter(contact_type='email').first()
        customer_email = email_contact.value if email_contact else ''
        
        # Get HR email (look for contact with 'HR' in label)
        hr_contact = customer.contacts.filter(contact_type='email', label__icontains='hr').first()
        hr_email = hr_contact.value if hr_contact else ''

        # Build address from customer fields
        street = customer.street_name or customer.address or ''
        house = f"{customer.house_number}{customer.house_number_addition}" if customer.house_number else ''
        if street and house:
            street_full = f"{street} {house}"
        else:
            street_full = street or house
        customer_address = f"{street_full}, {customer.postcode or ''} {customer.city or ''}".strip(', ')
        # Get customer logo path
        customer_logo_path = None
        if customer.logo and customer.logo.name:
            from django.conf import settings
            import os
            customer_logo_path = os.path.join(settings.MEDIA_ROOT, customer.logo.name)
    except Customer.DoesNotExist:
        return HttpResponse('Customer not found', status=404)


    # Parse week numbers
    week_from = None
    week_to = None
    if week_start:
        try:
            week_from = int(week_start.split('-W')[1])
        except (ValueError, IndexError):
            pass
    if week_end:
        try:
            week_to = int(week_end.split('-W')[1])
        except (ValueError, IndexError):
            pass
    
    # Build worklog queryset
    queryset = WorkEntry.objects.filter(
        project__customer_id=customer_id
    ).select_related(
        'employee', 'project', 'service', 'shift_template'
    )
    
    # Apply week filter
    if week_start or week_end:
        from django.db.models.functions import ExtractWeek, ExtractYear
        from django.db.models import Q
        
        if week_start:
            start_year, start_week = week_start.split('-W')
            if week_end:
                end_year, end_week = week_end.split('-W')
            else:
                end_year, end_week = start_year, start_week
            
            queryset = queryset.annotate(
                work_week=ExtractWeek('work_date'),
                work_year=ExtractYear('work_date')
            ).filter(
                Q(work_year=int(start_year), work_week__gte=int(start_week)) |
                Q(work_year__gt=int(start_year)),
                Q(work_year=int(end_year), work_week__lte=int(end_week)) |
                Q(work_year__lt=int(end_year))
            )
    
    # Apply employee filter
    if employee_ids and employee_ids[0]:
        queryset = queryset.filter(employee_id__in=employee_ids)
    
    # Apply supervisor filter - check both project.outfolder and project.supervisors
    if supervisor_id:
        from django.db.models import Q
        queryset = queryset.filter(
            Q(project__outfolder_id=supervisor_id) | 
            Q(project__supervisors__id=supervisor_id)
        ).distinct()

    
    # Convert to list of dicts for the export function
    worklogs = []
    for entry in queryset.order_by('work_date', 'actual_start_datetime'):
        # Get hours_breakdown from the model method
        hours_breakdown = {}
        if hasattr(entry, 'get_hours_breakdown_detailed'):
            try:
                hours_breakdown = entry.get_hours_breakdown_detailed()
            except Exception:
                pass  # Fall back to empty breakdown if calculation fails

        
        # Get start/end times - prefer actual, fallback to planned
        # Convert to Amsterdam local time for display
        from zoneinfo import ZoneInfo
        amsterdam_tz = ZoneInfo('Europe/Amsterdam')
        
        start_time = None
        end_time = None
        if entry.actual_start_datetime:
            local_dt = entry.actual_start_datetime.astimezone(amsterdam_tz) if entry.actual_start_datetime.tzinfo else entry.actual_start_datetime
            start_time = local_dt.strftime('%H:%M')
        elif entry.planned_start_time:
            start_time = entry.planned_start_time.strftime('%H:%M')
        
        if entry.actual_end_datetime:
            local_dt = entry.actual_end_datetime.astimezone(amsterdam_tz) if entry.actual_end_datetime.tzinfo else entry.actual_end_datetime
            end_time = local_dt.strftime('%H:%M')
        elif entry.planned_end_time:
            end_time = entry.planned_end_time.strftime('%H:%M')
        
        # Convert break duration minutes to HH:MM format
        # Use _get_total_break_minutes() which correctly sums breaks from JSON array
        break_mins = entry._get_total_break_minutes() if hasattr(entry, '_get_total_break_minutes') else (entry.break_duration_minutes or 0)
        break_duration = f"{break_mins // 60}:{break_mins % 60:02d}"

        
        # Extract break start/end times from breaks JSON array
        break_start = ''
        break_end = ''
        if entry.breaks and isinstance(entry.breaks, list) and len(entry.breaks) > 0:
            first_break = entry.breaks[0]
            if isinstance(first_break, dict):
                bs = first_break.get('start', '')
                be = first_break.get('end', '')
                # Format as HH:MM (remove seconds if present)
                if bs:
                    break_start = str(bs)[:5] if len(str(bs)) >= 5 else str(bs)
                if be:
                    break_end = str(be)[:5] if len(str(be)) >= 5 else str(be)
        
        worklogs.append({
            'employee_name': entry.employee.full_name if entry.employee else '',
            'project_name': entry.project.name if entry.project else '',
            'service_name': entry.service.name if entry.service else '',
            'work_date': entry.work_date,
            'start_time': start_time,
            'end_time': end_time,
            'break_duration': break_duration,
            'break_start': break_start,
            'break_end': break_end,
            'calculated_hours': entry.calculated_hours,
            'calculated_price': float(entry.calculated_price),  # Pre-calculated price with surcharges + allowances
            'hours_breakdown': hours_breakdown,
            'supervisor_name': entry.project.outfolder.company_name if entry.project and entry.project.outfolder else '',
            'hourly_rate': entry.get_service_rate() if hasattr(entry, 'get_service_rate') else 32.5,
            'allowances': entry.allowances if hasattr(entry, 'allowances') else [],
        })

    
    # Supervisor name for header
    rayon_name = None
    if supervisor_id:
        from apps.customers.models import Outfolder
        try:
            outfolder = Outfolder.objects.get(id=supervisor_id)
            rayon_name = outfolder.company_name
        except Outfolder.DoesNotExist:
            pass
    
    # Employee name for header (if single employee selected)
    medewerker_name = None
    if len(employee_ids) == 1 and employee_ids[0]:
        from apps.employees.models import EmployeeProfile
        try:
            emp = EmployeeProfile.objects.get(id=employee_ids[0])
            medewerker_name = emp.full_name
        except EmployeeProfile.DoesNotExist:
            pass
    
    # Get customer surcharges for dynamic percentage columns
    customer_surcharges = []
    for sc in customer.service_surcharges.filter(is_enabled=True).select_related('surcharge_type'):
        customer_surcharges.append({
            'name': sc.surcharge_type.name,
            'percentage': float(sc.percentage)
        })
    
    # Generate the Excel file
    buffer = generate_customer_excel_export(
        worklogs=worklogs,
        customer_name=customer_name,
        customer_email=customer_email,
        customer_address=customer_address,
        week_from=week_from,
        week_to=week_to,
        rayon_name=rayon_name,
        medewerker_name=medewerker_name,
        customer_logo_path=customer_logo_path,
        customer_surcharges=customer_surcharges,
        export_type=export_type,
        hr_email=hr_email
    )



    
    # Create HTTP response
    filename = f"Inleenovereenkomst_{customer_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
